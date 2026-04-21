#!/usr/bin/env python3
"""Post-process a sweep (w_sweep / m_sweep) into a section-vs-baseline Excel report.

Reads per_run.csv + summary.csv from the experiment directory and writes:
  comparison.xlsx — multi-sheet workbook (Summary, Comparison, PerRun)

The plain-text human-readable version is already produced by the sweep script
itself as summary.txt; this tool adds Excel output with a dedicated Comparison
sheet that has speedup / throughput_ratio columns for paper plots.

Usage: python3 make_comparison.py <experiment_dir>
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def detect_sweep_var(df: pd.DataFrame) -> str:
    for col in ("w", "m"):
        if col in df.columns:
            return col
    raise RuntimeError("cannot find sweep variable (w or m) in per_run.csv")


def build_comparison(summary: pd.DataFrame, var: str) -> pd.DataFrame:
    rows = []
    for x in sorted(summary[var].unique()):
        sec = summary[(summary["mode"] == "section") & (summary[var] == x)]
        base = summary[(summary["mode"] == "baseline") & (summary[var] == x)]
        if sec.empty or base.empty:
            continue
        s = sec.iloc[0]
        b = base.iloc[0]
        row = {var: x}
        for metric in [
            "throughput_per_step_mean",
            "sum_runtime_mean",
            "sum_hl_expanded_mean",
            "sum_runtime_detect_conf_mean",
            "sum_runtime_plan_paths_mean",
            "sum_runtime_find_consistent_mean",
            "system_success_count",
        ]:
            row[f"section_{metric}"] = s.get(metric)
            row[f"baseline_{metric}"] = b.get(metric)
        s_rt = s.get("sum_runtime_mean")
        b_rt = b.get("sum_runtime_mean")
        s_th = s.get("throughput_per_step_mean")
        b_th = b.get("throughput_per_step_mean")
        row["speedup"] = (b_rt / s_rt) if (s_rt and b_rt) else None
        row["throughput_ratio"] = (s_th / b_th) if (s_th and b_th) else None
        if "pure_section_count" in s.index:
            row["pure_section_count"] = s["pure_section_count"]
            row["total_fallback_events"] = s.get("total_fallback_events", 0)
        rows.append(row)
    return pd.DataFrame(rows)


def write_xlsx(out: Path, var: str, per_run: pd.DataFrame, summary: pd.DataFrame, comp: pd.DataFrame):
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        summary.sort_values(["mode", var]).to_excel(xw, sheet_name="Summary", index=False)
        comp.to_excel(xw, sheet_name="Comparison", index=False)
        per_run.to_excel(xw, sheet_name="PerRun", index=False)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="305496")
        center = Alignment(horizontal="center", vertical="center")
        for sheet in xw.sheets.values():
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
            for col_idx, col in enumerate(sheet.columns, start=1):
                max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 28)


def main():
    if len(sys.argv) != 2:
        print("usage: make_comparison.py <experiment_dir>", file=sys.stderr)
        sys.exit(2)
    exp_dir = Path(sys.argv[1])
    per_run_csv = exp_dir / "per_run.csv"
    summary_csv = exp_dir / "summary.csv"
    if not per_run_csv.exists() or not summary_csv.exists():
        print(f"missing per_run.csv or summary.csv in {exp_dir}", file=sys.stderr)
        sys.exit(1)

    per_run = pd.read_csv(per_run_csv)
    summary = pd.read_csv(summary_csv)
    var = detect_sweep_var(per_run)
    comp = build_comparison(summary, var)

    write_xlsx(exp_dir / "comparison.xlsx", var, per_run, summary, comp)
    print(f"wrote {exp_dir/'comparison.xlsx'}")


if __name__ == "__main__":
    main()
